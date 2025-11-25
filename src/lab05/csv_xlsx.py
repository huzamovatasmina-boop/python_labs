"""
Модуль для конвертации CSV в XLSX формат.
ЛР5 — JSON и конвертации
"""

import csv
from pathlib import Path

# Импортируем openpyxl (внешняя библиотека)
try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError("Для работы этого модуля установите openpyxl: pip install openpyxl")


def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    """
    Конвертирует CSV файл в XLSX формат (Excel).
    
    Автоматически подбирает ширину колонок по содержимому.
    
    Args:
        csv_path: Путь к исходному CSV файлу
        xlsx_path: Путь для сохранения XLSX файла
        
    Raises:
        FileNotFoundError: Если CSV файл не существует
        ValueError: Если CSV файл пустой
    """
    # Преобразуем пути в Path объекты
    csv_file = Path(csv_path)
    xlsx_file = Path(xlsx_path)
    
    # Проверяем существование CSV файла
    if not csv_file.exists():
        raise FileNotFoundError(f"CSV файл не найден: {csv_path}")
    
    # Читаем CSV файл
    with csv_file.open('r', encoding='utf-8') as cf:
        reader = csv.reader(cf)
        rows = list(reader)
    
    # Валидация данных
    if not rows:
        raise ValueError("CSV файл пустой")
    
    # Создаем родительские папки если их нет
    xlsx_file.parent.mkdir(parents=True, exist_ok=True)
    
    # Создаем новую книгу Excel
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    
    # Записываем данные из CSV в Excel
    for row in rows:
        sheet.append(row)
    
    # Настраиваем авто-ширину колонок
    for column_cells in sheet.columns:
        # Получаем максимальную длину текста в колонке
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        
        for cell in column_cells:
            try:
                # Проверяем длину текста в ячейке
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # Устанавливаем ширину колонки (минимум 8 символов)
        adjusted_width = max(max_length + 2, 8)
        sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Сохраняем XLSX файл
    workbook.save(xlsx_file)